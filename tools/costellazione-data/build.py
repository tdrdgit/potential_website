#!/usr/bin/env python3
"""Rigenera i sorgenti della costellazione a partire da un Excel.

    python3 tools/costellazione-data/build.py --bootstrap  # crea l'Excel dai json (una volta sola)
    python3 tools/costellazione-data/build.py --check      # non tocca niente, dice cosa cambierebbe
    python3 tools/costellazione-data/build.py              # riscrive i tre json nel Brain

Fonte unica dello strato editoriale della costellazione:
    costellazione/costellazione-dataset.xlsx   (gitignored: non esce su GitHub)

Scrive tre file nel Brain, gli stessi che scriveva l'editor visuale:
    suppliers database/taxonomy/tree.json    struttura: macro-ambiti, sotto-ambiti, settori, cartelle
    suppliers database/taxonomy/i18n.json    nome e descrizione di ogni nodo, quattro lingue
    suppliers database/taxonomy/scope.json   le quattro competenze di ogni nodo, quattro lingue

⚠️ L'Excel NON governa la dimensione dei nodi sul grafo. Il peso `w` lo calcola
build_public.py contando le cartelle dell'archivio DATABASE FORNITORI, e resta
un fatto misurato: se lo si potesse scrivere a mano smetterebbe di dire qualcosa.

⚠️ L'ordine delle righe del foglio NODI e' l'ordine dei nodi sul grafo. La
colonna `ord` esiste per renderlo un dato esplicito: se riordini il foglio per
lavorarci, riordina per `ord` prima di salvare, o la costellazione cambia
disposizione senza che nessuno l'abbia deciso.

⚠️ La lista dei sette settori NON e' qui e nemmeno in Excel: e' una costante di
taxonomy/build_taxonomy.py, perche' sono le categorie del sito. Il foglio
SETTORI ne governa le sole etichette tradotte, e questo script si ferma se i
suoi id non coincidono con i sette canonici.

Vale la catena di sempre. Dopo aver rigenerato i json:
    sh "…/suppliers database/scripts/publish.sh"   # rigenera e pubblica
    git pull                                       # il repo locale resta indietro di un commit
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Manca openpyxl.  Installalo con:  python3 -m pip install openpyxl")

REPO = Path(__file__).resolve().parents[2]
XLSX = REPO / "costellazione" / "costellazione-dataset.xlsx"

def find_drive() -> Path | None:
    """La cartella di Google Drive, senza scrivere l'indirizzo qui dentro.

    Il nome che monta Drive contiene l'indirizzo dell'account
    (GoogleDrive-<indirizzo>), e questo repository e' pubblico: scriverlo nel
    codice vorrebbe dire regalarlo al primo scraper che passa. Si cerca invece
    l'unica cartella che comincia per GoogleDrive-, che tra l'altro continua a
    funzionare se un giorno l'account cambia.
    """
    base = Path.home() / "Library" / "CloudStorage"
    if not base.exists():
        return None
    for folder in sorted(base.glob("GoogleDrive-*")):
        drive = folder / "Il mio Drive"
        if drive.exists():
            return drive
        drive = folder / "My Drive"
        if drive.exists():
            return drive
    return None


DRIVE = find_drive()
BRAIN = (DRIVE or Path.home() / "Google Drive") / "COMPANY BRAIN/POTENTIAL BRAIN/suppliers database"
TAX = BRAIN / "taxonomy"
BACKUP = TAX / ".backup"

TREE = TAX / "tree.json"
I18N = TAX / "i18n.json"
SCOPE = TAX / "scope.json"

# Le quattro lingue scritte nei sorgenti. Islandese e arabo sono tradotti per
# intero ma dal 2026-08-10 non escono nel file pubblico: restano di qua, e
# buttarli costerebbe una giornata per riaverli.
LANGS = ["it", "en", "is", "ar"]

# I sette Coverage Areas del sito. Copia di quelli di build_taxonomy.py, tenuta
# qui solo per controllare che il foglio SETTORI non se ne inventi uno: la
# lista buona resta quella, e cambiarla vuol dire cambiare il sito.
SECTORS = [
    "hospitality",
    "residential",
    "government",
    "retail",
    "prime-property",
    "yachting",
    "digital",
]

# La nota in testa a scope.json: e' documentazione per chi apre il file, non un
# dato che si edita in Excel, quindi vive qui e viene riscritta a ogni giro.
SCOPE_NOTE = (
    "Cosa apre ogni nodo quando lo si clicca: quattro voci, non di piu'. Sono "
    "raggruppamenti di capacita', non elenchi di materiali — 'Marmi e graniti "
    "da cava selezionata' dice al cliente cosa puo' chiedere, 'marmo, granito, "
    "ardesia, porfido' gli chiede di ricostruirselo. Vale la regola della "
    "sintesi del MoM: nominare poco e nominarlo bene. ⚠️ Il separatore e' il "
    "PUNTO E VIRGOLA, non la virgola: diverse voci contengono virgole interne "
    "('Cerniere, guide e sistemi di apertura') e l'arabo usa la propria "
    "virgola. Una voce separata male si spezza in due righe nella pagina. "
    "⚠️ Questo file lo genera tools/costellazione-data/build.py dall'Excel: "
    "editarlo a mano significa perderlo al primo rebuild."
)

SHEET_NODES = "NODI"
SHEET_UI = "UI"
SHEET_SECTORS = "SETTORI"
SHEET_CONFIG = "CONFIG"

# Intestazioni del foglio NODI, nell'ordine in cui compaiono. La chiave e' il
# nome della colonna in Excel, il valore la larghezza.
NODE_COLUMNS = [
    ("ord", 6),
    ("macro", 16),
    ("id", 22),
    ("sectors", 30),
    ("sources", 46),
]
for _lang in LANGS:
    NODE_COLUMNS.append((f"label_{_lang}", 26))
for _lang in LANGS:
    NODE_COLUMNS.append((f"blurb_{_lang}", 52))
for _lang in LANGS:
    NODE_COLUMNS.append((f"scope_{_lang}", 60))

HEAD_FILL = PatternFill("solid", fgColor="333333")
HEAD_FILL_MUTED = PatternFill("solid", fgColor="8A8A8A")
MACRO_FILL = PatternFill("solid", fgColor="EFEFEF")


# --------------------------------------------------------------------------- #
#  lettura dei json                                                            #
# --------------------------------------------------------------------------- #

def load_json(path: Path) -> dict:
    if not path.exists():
        sys.exit(f"manca {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def dump_json(payload: dict) -> str:
    """Stessa forma dei file scritti dall'editor: nessun diff cosmetico."""
    return json.dumps(payload, ensure_ascii=False, indent=1)


def split_cell(value: str | None) -> list[str]:
    """Una cella multi-valore -> lista.

    Separatore l'a capo dentro la cella (alt-invio), con il punto e virgola
    accettato come ripiego per chi incolla da altrove. Nessun nome di cartella
    dell'archivio contiene ne' l'uno ne' l'altro, quindi non c'e' ambiguita'.
    """
    if value is None:
        return []
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    parts = text.split("\n") if "\n" in text else text.split(";")
    return [p.strip() for p in parts if p.strip()]


def join_cell(values: list[str]) -> str:
    return "\n".join(values)


# --------------------------------------------------------------------------- #
#  bootstrap: json -> Excel                                                    #
# --------------------------------------------------------------------------- #

def bootstrap(force: bool) -> int:
    if XLSX.exists() and not force:
        sys.exit(
            f"{XLSX.name} esiste gia'.\n"
            "  Il bootstrap serve una volta sola, a creare l'Excel dai json. "
            "Rifarlo su un file gia' in uso ci scriverebbe sopra il contenuto "
            "del Brain, buttando le modifiche non ancora rigenerate.\n"
            "  Se e' davvero quello che vuoi:  --bootstrap --force"
        )

    tree = load_json(TREE)
    i18n = load_json(I18N)
    scope = load_json(SCOPE)

    wb = openpyxl.Workbook()

    # ---- foglio NODI ----
    ws = wb.active
    ws.title = SHEET_NODES
    ws.append([name for name, _ in NODE_COLUMNS])
    for index, (name, width) in enumerate(NODE_COLUMNS, start=1):
        letter = get_column_letter(index)
        ws.column_dimensions[letter].width = width
        cell = ws.cell(row=1, column=index)
        cell.font = Font(bold=True, color="FFFFFF")
        # Le colonne is/ar sono scritte ma non pubblicate: l'intestazione piu'
        # chiara lo dice a colpo d'occhio, senza bisogno di una legenda.
        muted = name.endswith("_is") or name.endswith("_ar")
        cell.fill = HEAD_FILL_MUTED if muted else HEAD_FILL
        cell.alignment = Alignment(vertical="center")

    order = 0
    for macro in tree["macros"]:
        macro_id = macro["id"]
        order += 1
        write_node_row(ws, order, macro_id, macro_id, None, None, i18n, scope, True)
        for child in macro["children"]:
            order += 1
            write_node_row(
                ws, order, macro_id, child["id"],
                child.get("sectors", []), child.get("sources", []),
                i18n, scope, False,
            )

    ws.freeze_panes = "D2"
    apply_outline(ws)

    # ---- foglio UI ----
    ws_ui = wb.create_sheet(SHEET_UI)
    write_lang_sheet(ws_ui, "key", i18n["ui"])

    # ---- foglio SETTORI ----
    ws_sec = wb.create_sheet(SHEET_SECTORS)
    write_lang_sheet(ws_sec, "id", i18n["sectors"])

    # ---- foglio CONFIG ----
    ws_cfg = wb.create_sheet(SHEET_CONFIG)
    ws_cfg.append(["chiave", "valore", "cos'e'"])
    for index in range(1, 4):
        cell = ws_cfg.cell(row=1, column=index)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
    ws_cfg.append([
        "version", tree.get("version", 1),
        "numero di versione del dataset, finisce in network.json",
    ])
    ws_cfg.append([
        "languages", ", ".join(i18n["languages"]),
        "le lingue che escono nel file pubblico. Le altre restano scritte e non escono.",
    ])
    ws_cfg.append([
        "rtl", ", ".join(i18n.get("rtl", [])),
        "lingue che si scrivono da destra a sinistra. Vuoto finche' l'arabo non esce.",
    ])
    ws_cfg.column_dimensions["A"].width = 14
    ws_cfg.column_dimensions["B"].width = 22
    ws_cfg.column_dimensions["C"].width = 80

    XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)

    leaves = sum(len(m["children"]) for m in tree["macros"])
    print(f"scritto {XLSX}")
    print(
        f"  {len(tree['macros'])} macro-ambiti, {leaves} sotto-ambiti, "
        f"{len(i18n['ui'])} stringhe di interfaccia, {len(SECTORS)} settori"
    )
    print("\nDa qui in avanti la fonte e' l'Excel. Per rigenerare i json:")
    print("  python3 tools/costellazione-data/build.py --check")
    return 0


def apply_outline(ws) -> tuple[int, int]:
    """Raggruppa le righe dei sotto-ambiti sotto il loro macro-ambito.

    E' la struttura a livelli di Excel: i numeri 1 e 2 in alto a sinistra
    chiudono e aprono tutto, il +/- accanto a ogni macro-ambito apre il suo
    ramo. Con 90 righe e' la differenza fra un elenco e una mappa.

    `summaryBelow = False` e' la riga che conta: dice a Excel che il capitolo
    sta SOPRA i suoi figli. Senza, il +/- finisce sulla riga sbagliata — sul
    macro-ambito successivo — e la struttura sembra rotta.

    Si ricalcola dalla colonna `macro`, non dalla posizione: una riga aggiunta
    in fondo al foglio prende il livello giusto lo stesso.
    """
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    header = [cell_text(c.value) for c in ws[1]]
    try:
        col_macro = header.index("macro")
        col_id = header.index("id")
    except ValueError:
        return (0, 0)

    grouped = 0
    for line in range(2, ws.max_row + 1):
        macro_id = cell_text(ws.cell(row=line, column=col_macro + 1).value)
        node_id = cell_text(ws.cell(row=line, column=col_id + 1).value)
        if not node_id:
            continue
        child = bool(macro_id) and macro_id != node_id
        ws.row_dimensions[line].outlineLevel = 1 if child else 0
        # Si nasce tutto aperto: chi vuole chiudere lo fa con un click, chi
        # apre un foglio a caso e lo trova chiuso pensa che manchino delle righe.
        ws.row_dimensions[line].hidden = False
        grouped += int(child)

    # Le colonne delle lingue che non escono nel file pubblico: tre coppie
    # (label, blurb, scope) che si chiudono in blocco e liberano meta' schermo.
    columns = 0
    for index, name in enumerate(header, start=1):
        if name and (name.endswith("_is") or name.endswith("_ar")):
            letter = get_column_letter(index)
            ws.column_dimensions[letter].outlineLevel = 1
            ws.column_dimensions[letter].hidden = False
            columns += 1

    return (grouped, columns)


def write_node_row(ws, order, macro_id, node_id, sectors, sources,
                   i18n, scope, is_macro) -> None:
    entry = i18n["nodes"].get(node_id, {})
    row = [
        order,
        macro_id,
        node_id,
        "; ".join(sectors) if sectors else None,
        join_cell(sources) if sources else None,
    ]
    for lang in LANGS:
        row.append(entry.get("label", {}).get(lang))
    for lang in LANGS:
        row.append(entry.get("blurb", {}).get(lang))
    for lang in LANGS:
        row.append(scope.get(node_id, {}).get(lang))
    ws.append(row)

    line = ws.max_row
    for index in range(1, len(NODE_COLUMNS) + 1):
        cell = ws.cell(row=line, column=index)
        cell.alignment = Alignment(vertical="top", wrap_text=index >= 4)
        if is_macro:
            # Le righe macro non portano ne' settori ne' cartelle: i settori di
            # un macro-ambito sono l'unione di quelli dei figli, e li calcola
            # build_public.py. Sfondo diverso perche' non le si compili per sbaglio.
            cell.fill = MACRO_FILL
            if index in (1, 2, 3):
                cell.font = Font(bold=True)


def write_lang_sheet(ws, key_header: str, entries: dict) -> None:
    ws.append([key_header] + LANGS)
    for index in range(1, len(LANGS) + 2):
        cell = ws.cell(row=1, column=index)
        cell.font = Font(bold=True, color="FFFFFF")
        muted = index > 3
        cell.fill = HEAD_FILL_MUTED if muted else HEAD_FILL
    for key, value in entries.items():
        ws.append([key] + [value.get(lang) for lang in LANGS])
    ws.column_dimensions["A"].width = 18
    for index in range(2, len(LANGS) + 2):
        ws.column_dimensions[get_column_letter(index)].width = 34
    ws.freeze_panes = "B2"


# --------------------------------------------------------------------------- #
#  lettura dell'Excel                                                          #
# --------------------------------------------------------------------------- #

def cell_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def read_workbook() -> dict:
    if not XLSX.exists():
        sys.exit(
            f"manca {XLSX}\n"
            "  e' l'Excel che governa la costellazione. Se non l'hai ancora "
            "creato:  python3 tools/costellazione-data/build.py --bootstrap"
        )
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    missing = [s for s in (SHEET_NODES, SHEET_UI, SHEET_SECTORS, SHEET_CONFIG)
               if s not in wb.sheetnames]
    if missing:
        sys.exit(
            "all'Excel mancano dei fogli: " + ", ".join(missing) + "\n"
            "  se sono stati rinominati, rimettili col nome di prima: lo script "
            "li cerca per nome, non per posizione."
        )

    ws = wb[SHEET_NODES]
    header = [cell_text(c.value) for c in ws[1]]
    expected = [name for name, _ in NODE_COLUMNS]
    if header[:len(expected)] != expected:
        sys.exit(
            "le intestazioni del foglio NODI non sono quelle attese.\n"
            f"  attese: {', '.join(expected)}\n"
            f"  trovate: {', '.join(str(h) for h in header)}\n"
            "  le colonne si leggono per nome: rinominarne una la rende invisibile."
        )
    index_of = {name: i for i, name in enumerate(header) if name}

    rows = []
    for line, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell_text(v) for v in raw):
            continue  # riga vuota: capita in coda al foglio, non e' un errore
        record = {
            "line": line,
            "ord": raw[index_of["ord"]] if index_of["ord"] < len(raw) else None,
            "macro": cell_text(raw[index_of["macro"]]),
            "id": cell_text(raw[index_of["id"]]),
            "sectors": split_cell(raw[index_of["sectors"]]),
            "sources": split_cell(raw[index_of["sources"]]),
        }
        for field in ("label", "blurb", "scope"):
            record[field] = {}
            for lang in LANGS:
                key = f"{field}_{lang}"
                position = index_of.get(key)
                value = cell_text(raw[position]) if position is not None and position < len(raw) else None
                if value:
                    record[field][lang] = value
        rows.append(record)

    return {
        "nodes": rows,
        "ui": read_lang_sheet(wb[SHEET_UI]),
        "sectors": read_lang_sheet(wb[SHEET_SECTORS]),
        "config": read_config(wb[SHEET_CONFIG]),
    }


def read_lang_sheet(ws) -> dict:
    header = [cell_text(c.value) for c in ws[1]]
    langs = [h for h in header[1:] if h]
    entries: dict[str, dict] = {}
    for raw in ws.iter_rows(min_row=2, values_only=True):
        key = cell_text(raw[0]) if raw else None
        if not key:
            continue
        entry = {}
        for offset, lang in enumerate(langs, start=1):
            value = cell_text(raw[offset]) if offset < len(raw) else None
            if value:
                entry[lang] = value
        entries[key] = entry
    return entries


def read_config(ws) -> dict:
    config: dict[str, str] = {}
    for raw in ws.iter_rows(min_row=2, values_only=True):
        key = cell_text(raw[0]) if raw else None
        if not key:
            continue
        config[key] = cell_text(raw[1]) if len(raw) > 1 else None
    return config


# --------------------------------------------------------------------------- #
#  validazione                                                                 #
# --------------------------------------------------------------------------- #

def validate(data: dict) -> list[str]:
    """Tutto quello che romperebbe la pubblicazione, detto in italiano.

    Si raccolgono TUTTI i problemi prima di uscire: chi sistema il foglio deve
    vedere l'elenco intero, non scoprirne uno per volta a ogni tentativo. Ogni
    riga porta il numero della riga di Excel, che e' l'unica coordinata utile
    quando si torna sul foglio.
    """
    problems: list[str] = []
    rows = data["nodes"]

    seen_ids: dict[str, int] = {}
    seen_sources: dict[str, str] = {}
    macro_ids = [r["id"] for r in rows if r["macro"] and r["id"] and r["macro"] == r["id"]]

    for row in rows:
        line = row["line"]
        node_id = row["id"]
        macro_id = row["macro"]

        if not node_id:
            problems.append(f"riga {line}: manca l'id del nodo")
            continue
        if not macro_id:
            problems.append(f"riga {line} ({node_id}): manca il macro-ambito di appartenenza")
            continue
        if node_id in seen_ids:
            problems.append(f"riga {line}: id \"{node_id}\" gia' usato alla riga {seen_ids[node_id]}")
        seen_ids[node_id] = line

        is_macro = macro_id == node_id
        if not is_macro and macro_id not in macro_ids:
            problems.append(
                f"riga {line} ({node_id}): il macro-ambito \"{macro_id}\" non esiste. "
                "Un macro-ambito e' una riga in cui le colonne macro e id sono uguali."
            )

        if is_macro:
            if row["sectors"]:
                problems.append(
                    f"riga {line} ({node_id}): un macro-ambito non porta settori — "
                    "sono l'unione di quelli dei suoi sotto-ambiti, li calcola build_public.py"
                )
            if row["sources"]:
                problems.append(
                    f"riga {line} ({node_id}): un macro-ambito non porta cartelle — "
                    "le cartelle si assegnano ai sotto-ambiti"
                )
        else:
            unknown = [s for s in row["sectors"] if s not in SECTORS]
            if unknown:
                problems.append(
                    f"riga {line} ({node_id}): settori sconosciuti {', '.join(unknown)}. "
                    f"I sette validi: {', '.join(SECTORS)}"
                )
            if not row["sectors"]:
                problems.append(f"riga {line} ({node_id}): nessun settore servito")
            if not row["sources"]:
                problems.append(
                    f"riga {line} ({node_id}): nessuna cartella dell'archivio assegnata — "
                    "il nodo esisterebbe sul grafo senza niente dietro"
                )
            for source in row["sources"]:
                if source in seen_sources:
                    problems.append(
                        f"riga {line}: cartella assegnata due volte: \"{source}\" "
                        f"({seen_sources[source]} e {node_id})"
                    )
                seen_sources[source] = node_id

        # Le lingue pubblicate vogliono tutto: una lingua su due rompe il selettore.
        published = [l.strip() for l in (data["config"].get("languages") or "it, en").split(",")]
        for field in ("label", "blurb", "scope"):
            missing = [l for l in published if not row[field].get(l)]
            if missing:
                problems.append(f"riga {line} ({node_id}): {field} manca in {', '.join(missing)}")

    for macro_id in macro_ids:
        children = [r for r in rows if r["macro"] == macro_id and r["id"] != macro_id]
        if not children:
            problems.append(f"{macro_id}: nessun sotto-ambito")

    # Il foglio SETTORI governa le etichette, non la lista: quella e' una
    # costante di build_taxonomy.py, e se le due divergono la pagina resta con
    # un filtro senza nome o con un nome senza filtro.
    sheet_sectors = list(data["sectors"].keys())
    if sheet_sectors != SECTORS:
        extra = [s for s in sheet_sectors if s not in SECTORS]
        gone = [s for s in SECTORS if s not in sheet_sectors]
        detail = []
        if extra:
            detail.append(f"in piu': {', '.join(extra)}")
        if gone:
            detail.append(f"mancanti: {', '.join(gone)}")
        if not detail:
            detail.append("stessi settori ma in ordine diverso")
        problems.append(
            "il foglio SETTORI non corrisponde ai sette Coverage Areas del sito "
            f"({'; '.join(detail)}). La lista buona sta in taxonomy/build_taxonomy.py: "
            "aggiungere un settore vuol dire cambiare il sito, non solo il foglio."
        )

    published = [l.strip() for l in (data["config"].get("languages") or "it, en").split(",")]
    for label, entries in (("UI", data["ui"]), ("SETTORI", data["sectors"])):
        for key, entry in entries.items():
            missing = [l for l in published if not entry.get(l)]
            if missing:
                problems.append(f"foglio {label}, {key}: manca {', '.join(missing)}")

    return problems


def warnings(data: dict) -> list[str]:
    """Cose storte che non fermano la pubblicazione ma vanno dette."""
    notes: list[str] = []

    for row in data["nodes"]:
        for lang in LANGS:
            text = row["scope"].get(lang)
            if text and len(text.split(";")) != 4:
                notes.append(
                    f"riga {row['line']} ({row['id']}): scope_{lang} ha "
                    f"{len(text.split(';'))} voci invece di 4 — il separatore e' "
                    "il punto e virgola, non la virgola"
                )
        published = [l.strip() for l in (data["config"].get("languages") or "it, en").split(",")]
        for field in ("label", "blurb", "scope"):
            missing = [l for l in LANGS if l not in published and not row[field].get(l)]
            if missing:
                notes.append(
                    f"riga {row['line']} ({row['id']}): {field} manca in "
                    f"{', '.join(missing)} — lingue non pubblicate, non blocca"
                )

    # Le cartelle dichiarate esistono davvero nell'archivio? Il controllo
    # definitivo lo fa build_public.py al momento di pubblicare, ma scoprirlo
    # qui costa niente e evita un giro a vuoto.
    suppliers = BRAIN / "data" / "suppliers.json"
    if suppliers.exists():
        try:
            known = {r.get("category") for r in json.loads(suppliers.read_text(encoding="utf-8"))}
        except (json.JSONDecodeError, TypeError, AttributeError):
            known = set()
        if known:
            for row in data["nodes"]:
                for source in row["sources"]:
                    if source not in known:
                        notes.append(
                            f"riga {row['line']} ({row['id']}): la cartella "
                            f"\"{source}\" non risulta nell'archivio scansionato — "
                            "controlla il nome, o riscansiona con refresh.py"
                        )

    return notes


# --------------------------------------------------------------------------- #
#  Excel -> json                                                               #
# --------------------------------------------------------------------------- #

def build_payloads(data: dict) -> dict[Path, str]:
    rows = data["nodes"]
    config = data["config"]

    version_raw = config.get("version") or "1"
    try:
        version = int(float(version_raw))
    except (TypeError, ValueError):
        version = version_raw
    languages = [l.strip() for l in (config.get("languages") or "it, en").split(",") if l.strip()]
    rtl = [l.strip() for l in (config.get("rtl") or "").split(",") if l.strip()]

    # L'ordine di apparizione delle righe e' l'ordine sul grafo: prima volta che
    # si incontra un macro-ambito, prima volta che si incontra un suo figlio.
    macros: list[dict] = []
    position: dict[str, int] = {}
    for row in rows:
        macro_id = row["macro"]
        if macro_id not in position:
            position[macro_id] = len(macros)
            macros.append({"id": macro_id, "children": []})
        if row["id"] != macro_id:
            macros[position[macro_id]]["children"].append({
                "id": row["id"],
                "sectors": row["sectors"],
                "sources": row["sources"],
            })

    tree = {"version": version, "sectors": SECTORS, "macros": macros}

    nodes_i18n = {}
    scope = {"_nota": SCOPE_NOTE}
    for row in rows:
        nodes_i18n[row["id"]] = {
            "label": {l: row["label"][l] for l in LANGS if l in row["label"]},
            "blurb": {l: row["blurb"][l] for l in LANGS if l in row["blurb"]},
        }
        scope[row["id"]] = {l: row["scope"][l] for l in LANGS if l in row["scope"]}

    i18n = {
        "languages": languages,
        "rtl": rtl,
        "ui": {k: {l: v[l] for l in LANGS if l in v} for k, v in data["ui"].items()},
        "sectors": {k: {l: v[l] for l in LANGS if l in v} for k, v in data["sectors"].items()},
        "nodes": nodes_i18n,
    }

    return {
        TREE: dump_json(tree),
        I18N: dump_json(i18n),
        SCOPE: dump_json(scope),
    }


def backup(paths: list[Path]) -> Path:
    """Stessa rete dell'editor: una copia datata prima di ogni sovrascrittura."""
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    folder = BACKUP / stamp
    folder.mkdir(parents=True, exist_ok=True)
    for path in paths:
        if path.exists():
            shutil.copy2(path, folder / path.name)
    return folder


def reoutline() -> int:
    """Rimette la struttura a livelli su un Excel gia' in uso.

    Serve dopo aver aggiunto righe: Excel eredita il livello di raggruppamento
    solo se la riga nasce dentro un gruppo, e una riga aggiunta in fondo resta
    fuori. Questo comando ricalcola tutto dalla colonna `macro` e non tocca il
    contenuto delle celle.

    ⚠️ Va lanciato a file chiuso: se l'Excel e' aperto, il salvataggio si scontra
    con la copia che Excel tiene in memoria e vince l'ultimo che scrive.
    """
    if not XLSX.exists():
        sys.exit(f"manca {XLSX}")
    wb = openpyxl.load_workbook(XLSX)
    if SHEET_NODES not in wb.sheetnames:
        sys.exit(f"all'Excel manca il foglio {SHEET_NODES}")
    rows, columns = apply_outline(wb[SHEET_NODES])
    wb.save(XLSX)
    print(f"raggruppate {rows} righe di sotto-ambiti e {columns} colonne (islandese e arabo)")
    print(f"  {XLSX.name}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--bootstrap", action="store_true",
                        help="crea l'Excel dai json esistenti, una volta sola")
    parser.add_argument("--force", action="store_true",
                        help="con --bootstrap: sovrascrive un Excel gia' esistente")
    parser.add_argument("--check", action="store_true",
                        help="non tocca niente, dice cosa cambierebbe")
    parser.add_argument("--outline", action="store_true",
                        help="riapplica il raggruppamento a righe e colonne dell'Excel")
    args = parser.parse_args()

    if args.outline:
        return reoutline()

    if not TAX.exists():
        sys.exit(
            f"non trovo la cartella dei sorgenti:\n  {TAX}\n"
            "  e' nel Brain su Google Drive: se Drive non e' montato, montalo "
            "e riprova. Se il path e' cambiato, va aggiornata la costante BRAIN "
            "in questo script."
        )

    if args.bootstrap:
        return bootstrap(args.force)

    data = read_workbook()

    problems = validate(data)
    if problems:
        sys.exit(
            "l'Excel della costellazione ha dei problemi — niente e' stato "
            "riscritto:\n  " + "\n  ".join(problems)
        )

    for note in warnings(data):
        print(f"⚠️  {note}")

    payloads = build_payloads(data)

    changed = []
    for path, content in payloads.items():
        before = path.read_text(encoding="utf-8") if path.exists() else None
        if before != content:
            changed.append(path)

    if not changed:
        print("i tre json sono gia' allineati all'Excel: niente da riscrivere.")
        return 0

    if args.check:
        print("cambierebbero:")
        for path in changed:
            print(f"  {path.name}")
        print("\nPer scrivere:  python3 tools/costellazione-data/build.py")
        return 0

    folder = backup(list(payloads.keys()))
    for path, content in payloads.items():
        path.write_text(content, encoding="utf-8")

    rows = data["nodes"]
    macros = sum(1 for r in rows if r["macro"] == r["id"])
    print(f"{macros} macro-ambiti, {len(rows) - macros} sotto-ambiti -> {TAX}")
    print(f"  riscritti: {', '.join(p.name for p in changed)}")
    print(f"  copia di prima in: {folder}")
    print("\nOra pubblica:")
    print(f'  sh "{BRAIN / "scripts" / "publish.sh"}"')
    print("  git pull        # il repo locale resta indietro di un commit")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
