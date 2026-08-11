#!/usr/bin/env python3
"""Rigenera l'elenco clienti del sito a partire dall'Excel dei progetti fornitori.

    python3 tools/portfolio-data/build.py            # riscrive le pagine
    python3 tools/portfolio-data/build.py --check    # non tocca niente, dice cosa cambierebbe

Fonte unica:
    ~/Dropbox/CARLO DOCUMENTI MAC/WORK - SUPPLIERS OFFERING/PT-SUPPLIERS PROJECTS LIST.xlsx
    foglio "SUPPLIERS PROJECTS"

Entrano solo le righe con WEBSITE = x. Le righe RISERVATO non escono mai da
quel file, nemmeno per sbaglio: vengono scartate anche se qualcuno le flagga.

I nomi dei fornitori non escono mai dall'Excel: la colonna FORNITORE diventa
una lettera (A al fornitore con piu' clienti) e nelle pagine finisce solo
quella. La legenda lettera -> nome vero si stampa a schermo e non si salva da
nessuna parte, perche' il repository e' pubblico.

Scrive dentro i blocchi delimitati da PORTFOLIO-DATA:START/END in:
    portfolio/potential-carousel.html      la griglia alfabetica dei clienti
    portfolio/potential-portfolio-map.html l'array CLIENTS del grafico radar

Tutto il resto delle due pagine non viene toccato.
"""

import argparse
import json
import re
import sys
from collections import OrderedDict, defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Manca openpyxl.  Installalo con:  python3 -m pip install openpyxl")

REPO = Path(__file__).resolve().parents[2]
XLSX = Path.home() / "Dropbox/CARLO DOCUMENTI MAC/WORK - SUPPLIERS OFFERING/PT-SUPPLIERS PROJECTS LIST.xlsx"
SHEET = "SUPPLIERS PROJECTS"
TIERS_FILE = Path(__file__).resolve().parent / "tiers.json"

CAROUSEL = REPO / "portfolio" / "potential-carousel.html"
MAP = REPO / "portfolio" / "potential-portfolio-map.html"

START = "PORTFOLIO-DATA:START"
END = "PORTFOLIO-DATA:END"

# SECTOR dell'Excel -> asse della portfolio map + etichetta mostrata nel dettaglio.
# Aggiungendo un settore nuovo in Excel va aggiunto qui, altrimenti lo script si ferma.
SECTORS = OrderedDict([
    ("hospitality & spa",       ("hospitality", "Hospitality & Spa")),
    ("hospitality & spa ",      ("hospitality", "Hospitality & Spa")),
    ("residential",             ("residential", "Residential")),
    ("residential & private",   ("residential", "Residential")),
    ("government buildings",    ("government",  "Government Buildings")),
    ("retail & luxury",         ("retail",      "Retail & Luxury")),
    ("retail",                  ("retail",      "Retail & Luxury")),
    ("prime property sourcing", ("prime",       "Prime Property Sourcing")),
    ("yachting & marine",       ("yachting",    "Yachting & Marine")),
    ("office & corporate",      ("office",      "Office & Corporate")),
])
# Righe che non sono progetti: si scartano senza far rumore.
SKIP_SECTORS = {"furniture product"}

# Località scritte in due modi nello stesso file: si normalizzano, altrimenti
# lo stesso posto compare due volte nella stessa riga di dettaglio.
PLACES = {
    "italia": "Italy",
    "parigi": "Paris",
    "londra": "London",
    "milano": "Milan",
    "torino": "Turin",
    "roma": "Rome",
    "harrods, londra": "Harrods, London",
    "venice fondaco": "Venezia Fondaco",
    "russia / turchia": "Russia / Turkey",
}


def clean(v):
    return str(v).strip() if v is not None else ""


def place(v):
    v = clean(v)
    return PLACES.get(v.casefold(), v)


def sort_key(name):
    return name.casefold()


def read_excel(path):
    """Righe pubblicabili dell'Excel, raggruppate per cliente."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Nel file non c'e' il foglio «{SHEET}» (ci sono: {', '.join(wb.sheetnames)})")
    ws = wb[SHEET]

    rows = list(ws.iter_rows(values_only=True))
    header = [clean(c).upper() for c in rows[0]]
    need = ["WEBSITE", "RISERVATO", "SECTOR", "FORNITORE", "CLIENT / PROJECT", "LOCATION", "CONTRACTOR"]
    missing = [c for c in need if c not in header]
    if missing:
        sys.exit("Nell'Excel mancano le colonne: " + ", ".join(missing))
    idx = {c: header.index(c) for c in need}

    clients, warnings = OrderedDict(), []
    for n, row in enumerate(rows[1:], start=2):
        get = lambda c: clean(row[idx[c]]) if idx[c] < len(row) else ""
        name = get("CLIENT / PROJECT")
        if not name:
            continue
        if get("RISERVATO"):
            continue
        if get("WEBSITE").casefold() != "x":
            continue

        sector = get("SECTOR")
        key = sector.casefold()
        if key in SKIP_SECTORS:
            continue
        if key not in SECTORS:
            warnings.append(f"riga {n}: settore «{sector}» sconosciuto ({name}) — riga saltata")
            continue
        axis, label = SECTORS[key]

        rec = clients.setdefault(name, {"axis": axis, "label": label, "places": [], "via": [], "sup": [], "rows": []})
        rec["rows"].append(n)
        if rec["axis"] != axis:
            warnings.append(f"{name}: compare sia in {rec['label']} sia in {label} — tengo {rec['label']}")
        loc = place(get("LOCATION"))
        if loc and loc not in rec["places"]:
            rec["places"].append(loc)
        via = clean(get("CONTRACTOR"))
        if via and via not in rec["via"]:
            rec["via"].append(via)
        sup = clean(get("FORNITORE"))
        if sup and sup not in rec["sup"]:
            rec["sup"].append(sup)

    out = []
    for name in sorted(clients, key=sort_key):
        rec = clients[name]
        if len(rec["via"]) > 1:
            warnings.append(f"{name}: piu' contractor ({' / '.join(rec['via'])}) — li scrivo tutti")
        detail = " · ".join([rec["label"]] + rec["places"])
        if rec["via"]:
            detail += " · for " + " / ".join(rec["via"])
        out.append({"name": name, "axis": rec["axis"], "detail": detail, "sup": rec["sup"], "rows": rec["rows"]})
    return out, warnings


def supplier_codes(clients):
    """Assegna una lettera a ogni fornitore: A a quello con piu' clienti.

    A parita' di clienti decide l'ordine alfabetico, cosi' la stessa tabella
    Excel produce sempre le stesse lettere. Il nome vero resta qui dentro: nelle
    pagine ci va solo la lettera, e chi guarda il sito non deve poterla
    ricondurre a nessuno.
    """
    counts = defaultdict(int)
    for c in clients:
        for s in c["sup"]:
            counts[s] += 1
    order = sorted(counts, key=lambda s: (-counts[s], sort_key(s)))
    if len(order) > 26:
        sys.exit(f"{len(order)} fornitori: le lettere dell'alfabeto non bastano piu'.")
    return {s: chr(65 + i) for i, s in enumerate(order)}, counts, order


def codes_of(client, codes):
    return "".join(sorted(codes[s] for s in client["sup"]))


def js(s):
    return s.replace("\\", "\\\\").replace("'", "\\'")


def html(s):
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def render_map(clients, tiers, codes):
    w = max(len(c["name"]) for c in clients) + 3
    ws = max(len(codes_of(c, codes)) for c in clients) + 6
    lines = ["const CLIENTS=["]
    for c in clients:
        n = f"'{js(c['name'])}',".ljust(w)
        a = f"a:'{c['axis']}',".ljust(18)
        s = f"s:'{codes_of(c, codes)}',".ljust(ws)
        lines.append(f" {{n:{n}{a}t:{tiers[c['name']]}, {s}d:'{js(c['detail'])}'}},")
    lines[-1] = lines[-1].rstrip(",")
    lines.append("];")
    return "\n".join(lines)


def render_carousel(clients):
    out = []
    for c in clients:
        out.append(f'    <div class="cc-alpha" data-detail="{html(c["detail"])}">')
        out.append(f'      <span class="cc-alpha-name">{html(c["name"])}</span>')
        out.append("    </div>")
        out.append("")
    return "\n".join(out).rstrip()


def splice(path, block, comment):
    """Sostituisce il testo fra i due marcatori. Se non ci sono, si ferma."""
    text = path.read_text(encoding="utf-8")
    pat = re.compile(r"(" + re.escape(START) + r".*?\n)(.*?)(\n[^\n]*" + re.escape(END) + r")", re.S)
    m = pat.search(text)
    if not m:
        sys.exit(f"In {path.name} mancano i marcatori {START} / {END}: non so dove scrivere.")
    new = text[:m.start(2)] + block + text[m.end(2):]
    return new, new != text


def current_names(path):
    text = path.read_text(encoding="utf-8")
    m = re.search(re.escape(START) + r".*?" + re.escape(END), text, re.S)
    if not m:
        return set()
    body = m.group(0)
    names = re.findall(r"cc-alpha-name\">(.*?)</span>", body)
    names += re.findall(r"\{n:'((?:[^'\\]|\\.)*)'", body)
    return {n.replace("&amp;", "&").replace("\\'", "'") for n in names}


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--check", action="store_true", help="non scrive niente, riporta solo le differenze")
    ap.add_argument("--xlsx", default=str(XLSX), help="percorso dell'Excel (default: quello di Carlo)")
    args = ap.parse_args()

    xlsx = Path(args.xlsx).expanduser()
    if not xlsx.exists():
        sys.exit(f"Non trovo l'Excel: {xlsx}")

    clients, warnings = read_excel(xlsx)
    if not clients:
        sys.exit("Nessuna riga con WEBSITE = x: non riscrivo le pagine con un elenco vuoto.")

    tiers_data = json.loads(TIERS_FILE.read_text(encoding="utf-8"))
    tiers = tiers_data["clients"]

    # Fasce mancanti: lo script non se le inventa, si ferma e le elenca.
    unknown = [c["name"] for c in clients if c["name"] not in tiers]
    if unknown:
        print("Clienti nuovi, senza fascia luxury assegnata:\n")
        for n in unknown:
            c = next(x for x in clients if x["name"] == n)
            print(f"  {n:<26} {c['detail']}")
        print(f"\nAssegna la fascia (1-5) in {TIERS_FILE.relative_to(REPO)} e rilancia.")
        print("1 mass market · 2 business standard · 3 premium · 4 alta gamma · 5 ultra luxury")
        sys.exit(1)

    before = current_names(MAP) | current_names(CAROUSEL)
    now = {c["name"] for c in clients}

    by_axis = defaultdict(int)
    for c in clients:
        by_axis[c["axis"]] += 1
    print(f"Excel: {xlsx.name}")
    print(f"{len(clients)} clienti pubblicabili — " + " · ".join(f"{a} {n}" for a, n in sorted(by_axis.items())))

    added, removed = sorted(now - before, key=sort_key), sorted(before - now, key=sort_key)
    if added:
        print("  + " + ", ".join(added))
    if removed:
        print("  − " + ", ".join(removed))
    if not added and not removed:
        print("  nessun cliente aggiunto o tolto (i dettagli possono comunque essere cambiati)")

    orphans = sorted(set(tiers) - now, key=sort_key)
    if orphans:
        print("  fasce senza cliente in Excel: " + ", ".join(orphans))

    codes, sup_counts, sup_order = supplier_codes(clients)
    print("\nFornitori — legenda che resta fra te e questo terminale:")
    for s in sup_order:
        print(f"  Supplier {codes[s]}  =  {s:<22} {sup_counts[s]} clienti")
    if not sup_order:
        print("  nessun fornitore compilato: la riga di filtri non compare")
    elif len(sup_order) == 1:
        print("  un fornitore solo: la riga di filtri resta nascosta finche' non sono almeno due")
    nosup = [c["name"] for c in clients if not c["sup"]]
    if nosup:
        warnings.append("senza FORNITORE in Excel: " + ", ".join(nosup)
                        + " — restano fuori da ogni filtro fornitore")
    for w in warnings:
        print("  ! " + w)

    jobs = [
        (MAP, render_map(clients, tiers, codes)),
        (CAROUSEL, render_carousel(clients)),
    ]
    changed = []
    for path, block in jobs:
        new, diff = splice(path, block, None)
        if diff:
            changed.append(path.name)
            if not args.check:
                path.write_text(new, encoding="utf-8")

    if args.check:
        print("\n--check: " + (", ".join(changed) + " da riscrivere" if changed else "le pagine sono gia' allineate"))
    else:
        print("\nRiscritte: " + (", ".join(changed) if changed else "nessuna, erano gia' allineate"))


if __name__ == "__main__":
    main()
